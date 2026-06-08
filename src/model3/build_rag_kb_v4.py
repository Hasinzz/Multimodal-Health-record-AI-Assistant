from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple


ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = ROOT / "data" / "improvement"
KB_DIR = ROOT / "data" / "rag_kb_v4"
OUT_DIR = ROOT / "outputs" / "v4_advanced_improvement" / "rag_kb_upgrade"


FIELD_ALIASES = {
    "medicine_name": ["Medicine Name", "Name", "product_name", "name"],
    "composition": ["Composition", "salt_composition", "short_composition1", "short_composition2"],
    "indication": ["Uses", "Indication", "medicine_desc", "Category", "sub_category"],
    "side_effects": ["Side_effects", "side_effects"],
    "interactions": ["drug_interactions"],
    "manufacturer": ["Manufacturer", "Manufacturer", "product_manufactured", "manufacturer_name"],
    "classification": ["Classification", "type"],
    "dosage_notes": ["Dosage Form", "Strength", "pack_size_label"],
}


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def first_value(row: Dict[str, str], columns: Iterable[str]) -> str:
    values = [clean_text(row.get(column, "")) for column in columns]
    return "; ".join(value for value in values if value)


def medicine_record_text(row: Dict[str, str], source: str, row_number: int) -> Tuple[str, str]:
    extracted = {key: first_value(row, columns) for key, columns in FIELD_ALIASES.items()}
    name = extracted["medicine_name"] or f"unknown medicine row {row_number}"
    parts = [f"Medicine: {name}."]
    if extracted["composition"]:
        parts.append(f"Composition: {extracted['composition']}.")
    if extracted["indication"]:
        parts.append(f"Indication or use: {extracted['indication']}.")
    if extracted["side_effects"]:
        parts.append(f"Side effects: {extracted['side_effects']}.")
    if extracted["interactions"]:
        parts.append(f"Interactions: {extracted['interactions']}.")
    if extracted["dosage_notes"]:
        parts.append(f"Dosage or pack notes: {extracted['dosage_notes']}.")
    if extracted["classification"]:
        parts.append(f"Classification: {extracted['classification']}.")
    if extracted["manufacturer"]:
        parts.append(f"Manufacturer: {extracted['manufacturer']}.")
    parts.append(f"Source: {source}, row {row_number}.")
    return name, " ".join(parts)


def iter_csv_records(path: Path, max_rows: int) -> Iterable[Tuple[int, Dict[str, str]]]:
    with path.open(newline="", encoding="utf-8", errors="ignore") as handle:
        reader = csv.DictReader(handle)
        for index, row in enumerate(reader, start=2):
            if index - 1 > max_rows:
                break
            yield index, row


def iter_xlsx_records(path: Path, max_rows: int) -> Iterable[Tuple[int, Dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as error:
        print(f"Skipping Excel file because openpyxl is unavailable: {path} ({error})")
        return

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [clean_text(value) for value in next(rows)]
    except StopIteration:
        workbook.close()
        return

    for count, values in enumerate(rows, start=2):
        if count - 1 > max_rows:
            break
        yield count, {header: clean_text(value) for header, value in zip(headers, values) if header}
    workbook.close()


def write_part(part_index: int, records: List[str]) -> Path:
    KB_DIR.mkdir(parents=True, exist_ok=True)
    path = KB_DIR / f"medicine_kb_v4_part_{part_index:03d}.txt"
    path.write_text("\n\n".join(records) + "\n", encoding="utf-8")
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build optional V4 medicine RAG KB text chunks.")
    parser.add_argument("--max-rows-per-file", type=int, default=3000)
    parser.add_argument("--records-per-kb-file", type=int, default=500)
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    KB_DIR.mkdir(parents=True, exist_ok=True)

    source_files = sorted((SOURCE_DIR / "model3_medicine_datasets").rglob("*.csv"))
    source_files.extend(sorted((SOURCE_DIR / "model3_medical_information").rglob("*.xlsx")))

    metadata_rows = []
    current_records: List[str] = []
    part_index = 1
    current_output = ""
    total_records = 0
    skipped = []

    for source_path in source_files:
        try:
            iterator = (
                iter_csv_records(source_path, args.max_rows_per_file)
                if source_path.suffix.lower() == ".csv"
                else iter_xlsx_records(source_path, args.max_rows_per_file)
            )
            for row_number, row in iterator or []:
                medicine_name, text = medicine_record_text(
                    row, str(source_path.relative_to(ROOT)), row_number
                )
                if len(text.split()) < 5:
                    continue
                current_records.append(text)
                total_records += 1
                if len(current_records) >= args.records_per_kb_file:
                    output_path = write_part(part_index, current_records)
                    current_output = str(output_path.relative_to(ROOT))
                    for item in current_records:
                        metadata_rows.append(
                            {
                                "kb_file": current_output,
                                "source_file": str(source_path.relative_to(ROOT)),
                                "medicine_name": item.split(".", 1)[0].replace("Medicine: ", ""),
                            }
                        )
                    current_records = []
                    part_index += 1
        except Exception as error:
            skipped.append(f"{source_path.relative_to(ROOT)}: {error}")

    if current_records:
        output_path = write_part(part_index, current_records)
        current_output = str(output_path.relative_to(ROOT))
        for item in current_records:
            metadata_rows.append(
                {
                    "kb_file": current_output,
                    "source_file": "mixed_final_part",
                    "medicine_name": item.split(".", 1)[0].replace("Medicine: ", ""),
                }
            )

    metadata_path = OUT_DIR / "rag_kb_v4_metadata.csv"
    with metadata_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["kb_file", "source_file", "medicine_name"])
        writer.writeheader()
        writer.writerows(metadata_rows)

    report_path = OUT_DIR / "rag_kb_v4_report.md"
    report_lines = [
        "# RAG KB V4 Report",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        f"Source files found: {len(source_files)}",
        f"Records converted: {total_records}",
        f"KB output folder: `{KB_DIR.relative_to(ROOT)}`",
        f"Metadata CSV: `{metadata_path.relative_to(ROOT)}`",
        "",
        "Existing `data/kb` files were not removed or modified.",
        "",
        "Skipped or limited files:",
    ]
    if skipped:
        report_lines.extend(f"- {item}" for item in skipped)
    else:
        report_lines.append("- None reported.")
    report_lines.extend(
        [
            "",
            "This KB is optional. Compare old KB versus V4 KB with medicine-related retrieval queries before making it the default.",
        ]
    )
    report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    print(f"Wrote KB files under {KB_DIR}")
    print(f"Wrote metadata: {metadata_path}")
    print(f"Wrote report: {report_path}")


if __name__ == "__main__":
    main()
